# ========================================================
#  
# Created by Stephen Covell on 21/12/21 @ 12:00 Hrs
#
# Project Title: Domain Discover
# Project Description:
#   To discover sub-domains and domains on a given domain.
#
# ========================================================

from openpyxl import Workbook
from datetime import datetime

# bad practise, but i need a quck work around
#wb = Workbook()

class Excel:
    """
    Creates and modifies excel documents using openpyxl.
    """    
        
    # Lets declare a workbook
    _wb = Workbook()

     # Lets declare a name for the file
    now = datetime.now()
    _wbname = now.strftime("saves/%d-%m-%Y %H-%M-%S.xlsx")

    def __init__(self):
        """
        __init___

        Sets the workbook information
        """

        # Lets create the required sheets
        self.createSheets()

        # Lets save the file
        self._wb.save(self._wbname)


    def createSheets(self):
        target_sheet = self.createTargetSheet()
        #vh_sheet = self.createVirtualHostSheet()

    def createTargetSheet(self):
        ws = self._wb.active
        ws.title = "Target"
        ws['A1'].value = 'Target'
        ws['B1'].value = 'Hostname'
        ws['C1'].value = 'IP'
        ws['D1'].value = 'Type'
        ws['E1'].value = 'Reverse DNS'
        ws['F1'].value = 'Netblock Owner'
        ws['G1'].value = 'Country'
        ws['H1'].value = 'Source'
        #ws['I1'].value = 'State'
        #ws['J1'].value = 'Ports'
        #ws['K1'].value = 'Extra Info'
        #ws['L1'].value = 'Reason'
        #ws['M1'].value = 'Version'
        #ws['N1'].value = 'Conf'
        return ws

    #def createNMapSheet(self):
    #    ws = wb.create_sheet("NMAP")
    #    ws = wb.active
    #    return ws
    
    #def openWB(self):
    #    wb.load_workbook('saves/test.xlsx')

    #def test(self):
    #    self.createSheets()
     #   wb.save('saves/test.xlsx')

    def addTargetSheetEntry(self, ws, targetdomain="", hostname="", ipaddr="", type="", reverseDNS="", netblockowner="", country="", source=""):
        ws.append([targetdomain, hostname, ipaddr, type, reverseDNS, netblockowner, country, source])
        self._wb.save(self._wbname)

#e = Excel()
#wse = e.createTargetSheet()
#e.addTargetSheetEntry(wse, "Target", "Hostname", "IP", "Type", "ReverseDNS", "Netblock", "Country", "Source3")


#wb = Workbook()
#ws = wb.active()

